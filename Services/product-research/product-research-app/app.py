"""
Product Research App
============================================
架构：
  用户提交主题 → 创建会话(pending) → 写入任务文件
  ↓
  cron 轮询任务文件 → 派发Agent
  ↓
  子Agent执行调研(Scrapling Skill + product-research Skill)
  ↓
  子Agent返回结构化结果 → 保存报告 → 标记done
"""
import json, os, sys, uuid, time, threading, traceback, re
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS

from io import BytesIO

app = Flask(__name__, static_folder='static', template_folder='templates', static_url_path='')
CORS(app)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONVOS_DIR = os.path.join(BASE_DIR, 'conversations')
TASKS_DIR = os.path.join(BASE_DIR, 'tasks')
WORKER_LOG = os.path.join(BASE_DIR, 'worker.log')
os.makedirs(CONVOS_DIR, exist_ok=True)
os.makedirs(TASKS_DIR, exist_ok=True)

# ============================================================
# Excel 生成
# ============================================================
def _clean_ingredients(raw):
    """将 Supplement Facts 原文精简为 '成分名 剂量; ...' 格式"""
    import re
    if not raw or raw == 'unknown':
        return raw
    lines = [l.strip() for l in raw.split('\n') if l.strip()]
    # 找到表头结束位置
    start = 0
    for i, l in enumerate(lines):
        if '% daily value' in l.lower() or 'daily value' in l.lower().replace('not established', ''):
            if 'not established' not in l.lower():
                start = i + 1
                break
    results = []
    i = start
    while i < len(lines):
        l = lines[i]
        if 'daily value not' in l.lower() or l.startswith('*'):
            break
        # 跳过百分比行
        if re.match(r'^[\d.]+%$|^\*+$', l):
            i += 1
            continue
        # 剂量行（纯数字+单位）不是成分名
        if re.match(r'^\d[\d,.]*\s*(mg|mcg|µg|g|IU|CFU|billion)', l):
            i += 1
            continue
        # 当前行是成分名，下一行可能是剂量
        if i + 1 < len(lines) and re.match(r'^\d[\d,.]*\s*(mg|mcg|µg|g|IU|CFU|billion)', lines[i + 1]):
            results.append(f"{l} {lines[i+1]}")
            i += 3  # 跳过成分、剂量、百分比
        else:
            i += 1
    return '; '.join(results) if results else raw

def generate_excel(output, report_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    hf, hs = Font(bold=True,color="FFFFFF",size=11), PatternFill("solid",fgColor="2F5496")
    bt = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
    ca, wrap = Alignment(horizontal='center',vertical='center',wrap_text=True), Alignment(vertical='center',wrap_text=True)

    ws1 = wb.active; ws1.title = "调研概览"
    ws1.merge_cells('A1:F1'); ws1['A1'] = f"📊 市场调研报告 — {output['研究范围']}"; ws1['A1'].font = Font(bold=True,size=16,color="2F5496")
    r = 3; ws1[f'A{r}']="研究范围"; ws1[f'A{r}'].font=Font(bold=True); ws1.merge_cells(f'B{r}:F{r}'); ws1[f'B{r}']=output["研究范围"]; ws1[f'B{r}'].alignment=wrap; r+=1
    ws1[f'A{r}']="📊 聚合统计"; ws1[f'A{r}'].font=Font(bold=True,size=13); r+=1
    for k,v in output["结构化结果"]["聚合统计"].items():
        ws1[f'A{r}']=k; ws1[f'A{r}'].font=Font(bold=True); ws1[f'B{r}']=str(v) if not isinstance(v,list) else ", ".join(map(str,v)); r+=1
    r+=1; ws1[f'A{r}']="✅ 关键结论"; ws1[f'A{r}'].font=Font(bold=True,size=13); r+=1
    for c in output.get("关键结论",[]):
        ws1[f'A{r}']=c; ws1.merge_cells(f'A{r}:F{r}'); ws1[f'A{r}'].alignment=wrap; r+=1
    r+=1; ws1[f'A{r}']="⚠️ 不确定性与缺口"; ws1[f'A{r}'].font=Font(bold=True,size=13); r+=1
    for g in output.get("不确定性与缺口",[]):
        ws1[f'A{r}']=g; ws1.merge_cells(f'A{r}:F{r}'); ws1[f'A{r}'].alignment=wrap; r+=1
    ws1.column_dimensions['A'].width=22; ws1.column_dimensions['B'].width=65

    ws2 = wb.create_sheet(title="产品明细")
    hdrs = ["序号","产品名称","品牌","平台来源","产品链接","剂型","规格","价格","核心卖点","核心原料","宣传方向","热度信号","适用人群"]
    for col,h in enumerate(hdrs,1):
        c = ws2.cell(row=1,column=col,value=h); c.font=hf; c.fill=hs; c.alignment=ca; c.border=bt
    for ri,rec in enumerate(output["结构化结果"]["records"], start=2):
        def to_cell(v):
            if isinstance(v, list):
                return ", ".join(str(x) for x in v)
            return v
        vals = [ri-1, to_cell(rec.get("product_name","")), to_cell(rec.get("brand","")), to_cell(rec.get("source_platform","")),
               to_cell(rec.get("product_url","")), to_cell(rec.get("dosage_form","")), to_cell(rec.get("pack_size","")),
               to_cell(rec.get("price","")), to_cell(rec.get("core_selling_points","")), to_cell(rec.get("core_ingredients","")),
               to_cell(rec.get("claim_direction","")), to_cell(rec.get("public_heat_signal","")), to_cell(rec.get("target_population",""))]
        for col,val in enumerate(vals,1):
            c = ws2.cell(row=ri,column=col,value=val); c.alignment=wrap; c.border=bt
    for i,w in enumerate([6,35,20,22,45,15,15,12,35,30,25,25,22],1): ws2.column_dimensions[get_column_letter(i)].width=w
    ws2.freeze_panes = 'A2'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ============================================================
# 对话文件管理
# ============================================================
def get_conversations():
    convos = []
    for f in sorted(os.listdir(CONVOS_DIR), reverse=True):
        if f.endswith('.json'):
            with open(os.path.join(CONVOS_DIR,f),'r',encoding='utf-8') as fh:
                d = json.load(fh)
            convos.append({'id':f.replace('.json',''),'title':d.get('title','Untitled'),
                         'created':d.get('created',''),'updated':d.get('updated',''),
                         'status':d.get('status','pending'),'report_id':d.get('report_id',''),
                         'topic': d.get('topic','')})
    return convos

def load_convo(cid):
    p = os.path.join(CONVOS_DIR, f'{cid}.json')
    return json.load(open(p,'r',encoding='utf-8')) if os.path.exists(p) else None

def save_convo(cid, data):
    json.dump(data, open(os.path.join(CONVOS_DIR,f'{cid}.json'),'w',encoding='utf-8'), ensure_ascii=False, indent=2)

def _log(msg):
    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    line = f"[{ts}] {msg}"
    print(line)
    with open(WORKER_LOG, 'a', encoding='utf-8') as f:
        f.write(line + '\n')

# ============================================================
# Schema 辅助函数
# ============================================================
def make_aggregation(records):
    """生成聚合统计数据"""
    ing_set, df_map, sp_map, products = set(), {}, {}, []
    pb = {"<50":0,"50-100":0,"100-200":0,"200-500":0,">500":0}
    platforms = set()

    for r in records:
        # 提取原料
        for i in (r.get("core_ingredients") or "").replace("；",";").replace(",", ";").split(";"):
            if (i:=i.strip()) and i not in ["N/A(demo)", "unknown", ""]:
                ing_set.add(i)
        # 统计剂型
        d = r.get("dosage_form","")
        if d and d not in ["unknown", "N/A(demo)", ""]:
            df_map[d] = df_map.get(d,0)+1
        # 统计卖点
        for s in (r.get("core_selling_points") or "").replace("；",";").split(";"):
            if (s:=s.strip()) and s not in ["N/A(demo)", ""]:
                sp_map[s] = sp_map.get(s,0)+1
        # 产品名称
        pn = r.get("product_name","")
        if pn and pn not in ["unknown", "", "N/A(demo)"]:
            products.append(pn)
        # 平台
        if r.get("source_platform"):
            platforms.add(r["source_platform"])
        # 价格带统计
        price = r.get("price", "")
        if isinstance(price, str):
            nums = re.findall(r'[\d.]+', price.replace(",", ""))
            if nums:
                try:
                    p = float(nums[0])
                    if p < 50: pb["<50"] += 1
                    elif p < 100: pb["50-100"] += 1
                    elif p < 200: pb["100-200"] += 1
                    elif p < 500: pb["200-500"] += 1
                    else: pb[">500"] += 1
                except: pass

    return {
        "total_product_count": len(records),
        "top_ingredients": list(ing_set)[:15] or ["待采集"],
        "top_dosage_forms": [f"{k}({v})" for k,v in sorted(df_map.items(),key=lambda x:-x[1])[:8]] or ["待采集"],
        "top_selling_points": [f"{k}({v})" for k,v in sorted(sp_map.items(),key=lambda x:-x[1])[:10]] or ["待采集"],
        "price_band_distribution": pb,
        "hot_selling_products": products[:15] or ["待采集"],
        "platform_differences": list(platforms) or ["待采集"],
    }

# ============================================================
# 子Agent进度保存/恢复（断点续传）
# ============================================================
# 子Agent逐条记录推送 + 前端增量拉取
# ============================================================
def _records_path(cid):
    return os.path.join(TASKS_DIR, f'{cid}_records.json')

def _load_records(cid):
    p = _records_path(cid)
    if os.path.exists(p):
        with open(p, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def _save_records(cid, records):
    with open(_records_path(cid), 'w', encoding='utf-8') as f:
        json.dump(records, f, ensure_ascii=False)

# 提交频率追踪：检测同一 source_html_path 短时间内大量提交（硬编码列表造假特征）
_submit_log = {}  # {cid: [(timestamp, source_html_path), ...]}

@app.route('/api/subagent/record/<cid>', methods=['POST'])
def append_record(cid):
    """子Agent每采集到一条产品记录就POST过来，后端追加存储"""
    try:
        rec = (request.json or {}).get('record')
        if not rec or not isinstance(rec, dict):
            return jsonify({'ok': False, 'error': 'Missing record'}), 400

        # ── 防造假校验：source_html_path 必须存在且内容可验证 ──
        html_path = rec.get('source_html_path', '')
        if not html_path:
            return jsonify({'ok': False, 'error': 'Missing source_html_path — every record must link to a scraped HTML file'}), 400
        if not os.path.isfile(html_path):
            return jsonify({'ok': False, 'error': f'source_html_path not found: {html_path}'}), 400
        # 文件不能太小（空壳/Cloudflare 拦截页通常 < 2KB）
        fsize = os.path.getsize(html_path)
        if fsize < 2000:
            return jsonify({'ok': False, 'error': f'source HTML too small ({fsize}B) — likely blocked or empty'}), 400
        # 产品名关键词必须出现在 HTML 中（防止随意指定不相关文件）
        pn = rec.get('product_name', '')
        with open(html_path, 'r', encoding='utf-8', errors='replace') as fh:
            html_sample = fh.read(500_000)  # 读前 500KB 足够
        # 取产品名中长度 > 2 的词做模糊匹配
        keywords = [w for w in re.split(r'[\s\-\+\(\)]+', pn) if len(w) > 2]
        if keywords and not any(kw.lower() in html_sample.lower() for kw in keywords):
            return jsonify({'ok': False, 'error': f'product_name keywords not found in source HTML — data may be fabricated'}), 400

        # 数据质量过滤
        if '_' in pn and ('推荐' in pn or '搜索' in pn) and any(k in pn for k in ['苏宁','京东','天猫','淘宝']):
            return jsonify({'ok': False, 'error': f'Rejected: search page title, not product name'}), 400
        if all(rec.get(k) in ['unknown','',None] for k in ['brand','price','core_ingredients']):
            return jsonify({'ok': False, 'error': 'Rejected: brand+price+ingredients all unknown'}), 400

        # ── unknown 字段过多：内容字段超半数为 unknown 则拒绝 ──
        _content_fields = ['product_name','brand','dosage_form','pack_size','price',
                           'core_selling_points','core_ingredients','claim_direction',
                           'public_heat_signal','target_population']
        unknown_cnt = sum(1 for f in _content_fields if rec.get(f) in ['unknown','',None])
        if unknown_cnt >= 7:
            return jsonify({'ok': False, 'error': f'Rejected: {unknown_cnt}/{len(_content_fields)} content fields are unknown — too little useful data'}), 400

        # ── 详情页链接失效检测：product_url 明显无效则拒绝 ──
        url = rec.get('product_url', '')
        if not url or url in ['unknown', 'N/A', 'none', 'null']:
            return jsonify({'ok': False, 'error': 'Rejected: product_url is missing or invalid'}), 400

        records = _load_records(cid)
        if url and any(r.get('product_url') == url for r in records):
            return jsonify({'ok': False, 'error': 'Duplicate product_url'}), 400
        # 禁止搜索 URL 作为 product_url（防止伪造来源）
        if url and any(marker in url for marker in ['?k=', '/s?', '/search?', '/search/']):
            return jsonify({'ok': False, 'error': 'product_url is a search URL, not a product page — use /dp/ or /pr/ URLs'}), 400

        # 频率检测：同一 source_html_path 在 10 秒内提交 5+ 条 → 疑似硬编码列表批量灌入
        now = time.time()
        log = _submit_log.setdefault(cid, [])
        log.append((now, html_path))
        recent_same = sum(1 for ts, p in log if p == html_path and now - ts < 10)
        if recent_same > 5:
            _log(f"⚠️ 频率警告 {cid}: {html_path} 在10秒内提交了 {recent_same} 条，疑似硬编码批量提交")

        records.append(rec)
        _save_records(cid, records)
        # 更新会话的 record_count
        c = load_convo(cid)
        if c:
            if c.get('status') == 'pending':
                c['status'] = 'running'
                c['started_at'] = datetime.now().isoformat()
            c['record_count'] = len(records)
            save_convo(cid, c)
        return jsonify({'ok': True, 'count': len(records)})
    except Exception as e:
        _log(f"❌ 追加记录失败 {cid}: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/conversation/<cid>/records')
def get_records(cid):
    """前端增量拉取：?offset=N 返回第N条之后的新记录"""
    offset = request.args.get('offset', 0, type=int)
    records = _load_records(cid)
    return jsonify({'records': records[offset:], 'total': len(records)})

# ============================================================
# API Routes
# ============================================================

@app.route('/')
def index(): return render_template('index.html')

@app.route('/api/conversations', methods=['GET'])
def api_convos(): return jsonify(get_conversations())

@app.route('/api/conversation/<cid>/status', methods=['GET'])
def api_status(cid):
    c = load_convo(cid)
    if not c: return jsonify({'error':'Not found'}), 404
    return jsonify({
        'id': cid, 'title': c.get('title',''), 'status': c.get('status','unknown'),
        'topic': c.get('topic',''), 'created': c.get('created',''),
        'started_at': c.get('started_at',''), 'completed_at': c.get('completed_at',''),
        'report_id': c.get('report_id'), 'record_count': c.get('record_count'),
        'error': c.get('error'),
    })

@app.route('/api/submit', methods=['POST'])
def api_submit():
    """提交研究主题 → 创建会话(pending) → 写入任务文件等待Agent处理"""
    data = request.json or {}
    topic = (data.get('topic') or data.get('message') or '').strip()
    if not topic: return jsonify({'error':'请输入研究主题'}), 400

    cid = uuid.uuid4().hex[:8]
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    convo = {
        'id': cid,
        'title': topic[:50] + ('...' if len(topic)>50 else ''),
        'topic': topic,
        'created': now,
        'updated': now,
        'status': 'pending',
        'report_id': None,
    }
    save_convo(cid, convo)

    task_file = os.path.join(TASKS_DIR, f'{cid}.json')
    task_def = {
        'cid': cid,
        'topic': topic,
        'status': 'pending',
        'created_at': datetime.now().isoformat()
    }
    with open(task_file, 'w', encoding='utf-8') as f:
        json.dump(task_def, f, ensure_ascii=False, indent=2)

    _log(f"📝 新提交: {cid} | {topic}")
    _log(f"📋 任务文件已写入: {task_file}")
    _log(f"⏳ 等待Agent处理...")

    return jsonify({'conversation_id': cid, 'status': 'pending'})

@app.route('/api/worker', methods=['POST'])
def api_worker():
    """手动触发 Worker：调用 cron_gateway_worker.py 扫描并派发 pending 任务"""
    import subprocess
    worker_script = os.path.join(BASE_DIR, 'cron_gateway_worker.py')
    try:
        result = subprocess.run(
            [sys.executable, worker_script],
            capture_output=True, text=True, timeout=60
        )
        _log(f"Worker 执行结果: {result.stdout[-500:]}")
        return jsonify({'ok': True, 'ts': datetime.now().isoformat(), 'message': 'Worker 已执行'})
    except Exception as e:
        _log(f"Worker 执行失败: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/conversation/<cid>/cancel', methods=['POST'])
def api_cancel(cid):
    """取消正在进行的调研"""
    c = load_convo(cid)
    if not c:
        return jsonify({'ok': False, 'error': '会话不存在'}), 404

    if c.get('status') in ['done', 'error']:
        return jsonify({'ok': False, 'error': f'会话已{c["status"]}，无法取消'}), 400

    # 更新状态为 cancelled
    c['status'] = 'error'
    c['error'] = '用户手动终止调研'
    c['updated'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    save_convo(cid, c)
    _log(f"⏹ 用户终止调研: {cid}")

    return jsonify({'ok': True, 'message': '调研已终止'})

@app.route('/api/conversation/<cid>/reset', methods=['POST'])
def api_reset(cid):
    """重置会话状态为 pending，允许重新提交处理"""
    c = load_convo(cid)
    if not c:
        return jsonify({'ok': False, 'error': '会话不存在'}), 404

    c['status'] = 'pending'
    c['error'] = None
    c['updated'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    if 'started_at' in c: del c['started_at']
    if 'completed_at' in c: del c['completed_at']
    if 'report_id' in c: c['report_id'] = None
    if 'record_count' in c: c['record_count'] = None
    save_convo(cid, c)

    # 清理逐条记录临时文件
    rp = _records_path(cid)
    if os.path.exists(rp):
        os.remove(rp)

    # 同步重置任务文件状态，否则 cron_worker 无法重新拾取
    task_file = os.path.join(TASKS_DIR, f'{cid}.json')
    if os.path.exists(task_file):
        with open(task_file, 'r', encoding='utf-8') as f:
            task = json.load(f)
        task['status'] = 'pending'
        for k in ('dispatched_at',):
            task.pop(k, None)
        with open(task_file, 'w', encoding='utf-8') as f:
            json.dump(task, f, ensure_ascii=False, indent=2)

    _log(f"🔄 重置会话: {cid} -> pending")

    return jsonify({'ok': True, 'message': '状态已重置'})

@app.route('/api/conversation/<cid>/excel')
def dl_excel(cid):
    """按需从当前已有记录生成 Excel 并下载"""
    records = _load_records(cid)
    if not records:
        return jsonify({'error': 'No records yet'}), 404
    topic = (load_convo(cid) or {}).get('topic', '')
    output = {
        "研究范围": topic,
        "结构化结果": {"records": records, "聚合统计": make_aggregation(records)},
        "关键结论": [], "不确定性与缺口": [],
    }
    buf = generate_excel(output, cid)
    return send_file(buf, as_attachment=True, download_name=f'market_research_{cid}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/conversation/<cid>/json')
def dl_json(cid):
    """按需从当前已有记录生成 JSON 并下载"""
    records = _load_records(cid)
    if not records:
        return jsonify({'error': 'No records yet'}), 404
    topic = (load_convo(cid) or {}).get('topic', '')
    output = {
        "研究范围": topic,
        "结构化结果": {"records": records, "聚合统计": make_aggregation(records)},
        "关键结论": [], "不确定性与缺口": [],
    }
    return jsonify(output)

# ============================================================
# 清空所有记录
# ============================================================
@app.route('/api/clear-all', methods=['POST'])
def api_clear_all():
    """清空所有调研记录和数据文件"""
    deleted = 0

    try:
        # 清空 conversations
        for f in os.listdir(CONVOS_DIR):
            if f.endswith('.json'):
                os.remove(os.path.join(CONVOS_DIR, f))
                deleted += 1

        # 清空 tasks（记录文件 + 任务文件）
        for f in os.listdir(TASKS_DIR):
            fpath = os.path.join(TASKS_DIR, f)
            if os.path.isfile(fpath):
                os.remove(fpath)

        _log(f"🗑️ 清空所有记录: {deleted} 条会话")
        return jsonify({'ok': True, 'deleted': deleted})

    except Exception as e:
        _log(f"❌ 清空失败: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500

if __name__ == '__main__':
    print("=" * 50)
    print("🦞 Product Research App")
    print(f"   http://127.0.0.1:5001")
    print("=" * 50)
    app.run(host='127.0.0.1', port=5001, debug=False)
